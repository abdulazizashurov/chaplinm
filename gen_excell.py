from history import History
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from db_user import DBUser
class GenExcell():
    """Класс генерации excell отчета"""
    @staticmethod
    def gen_history(file_name : str, date_from : str, date_to : str) -> None:
        """Генерация истории"""
        wb = Workbook()
        ws = wb.active
        ws.sheet_properties.pageSetUpPr.fitToPage = True    
        fields = ["Имя клиента", "Заказ", "дата заказа", "Цена", "Метод оплаты", "Тип доставки"]
        fields_width = [22.57, 37.71, 16.86, 11.71, 15.57, 17]
        for i in range(65, 71):
            cell = ws[f"{chr(i)}1"]
            cell.alignment = Alignment("center", "center")
            cell.value = fields[i-65]
            ws.column_dimensions[chr(i)].width = fields_width[i-65]
        pos = 2
        for history in History.get_histories(date_from, date_to):
            order_text = ""
            for i in range(65, 71):
                order_text = history[i-65]
                cell = ws[f"{chr(i)}{pos}"]
                cell.alignment = Alignment("left", "center", wrap_text=True)
                cell.value = order_text
            pos += 1
        wb.save(file_name)
        wb.close()
        wb = None
    @staticmethod
    def gen_users_list_Who_not_buy(file_name : str) -> None:
        """Экспорт данных пользователей кто не совершал покупки"""
        wb = Workbook()
        ws = wb.active
        ws.sheet_properties.pageSetUpPr.fitToPage = True    
        fields = ["Имя клиента", "юзернейм", "Номер телефона", "Язык", "Дата регистрации", 
                "Дата последней активности", "Кешбек", "Вся закупка", "Закупка за месяц", "Количество заказа"]
        fields_width = [22.57, 37.71, 16.86, 11.71, 15.57, 17, 15.57, 15.57, 17, 17]
        for i in range(65, 75):
            cell = ws[f"{chr(i)}1"]
            cell.alignment = Alignment("center", "center")
            cell.value = fields[i-65]
            ws.column_dimensions[chr(i)].width = fields_width[i-65]
        pos = 2
        for user in DBUser.load_all_users():
            if user[10] == 0:
                for i in range(65, 75):
                    cell = ws[f"{chr(i)}{pos}"]
                    cell.alignment = Alignment("left", "center", wrap_text=True)
                    cell.value = user[i-64]
                pos += 1
        wb.save(file_name)
        wb.close()
        wb = None
    @staticmethod
    def gen_users_list(file_name : str) -> None:
        """Экспорт данных пользователей"""
        wb = Workbook()
        ws = wb.active
        ws.sheet_properties.pageSetUpPr.fitToPage = True    
        fields = ["Имя клиента", "юзернейм", "Номер телефона", "Язык", "Дата регистрации", 
                "Дата последней активности", "Кешбек", "Вся закупка", "Закупка за месяц", "Количество заказа"]
        fields_width = [22.57, 37.71, 16.86, 11.71, 15.57, 17, 15.57, 15.57, 17, 17]
        for i in range(65, 75):
            cell = ws[f"{chr(i)}1"]
            cell.alignment = Alignment("center", "center")
            cell.value = fields[i-65]
            ws.column_dimensions[chr(i)].width = fields_width[i-65]
        pos = 2
        for user in DBUser.load_all_users():
            for i in range(65, 75):
                cell = ws[f"{chr(i)}{pos}"]
                cell.alignment = Alignment("left", "center", wrap_text=True)
                cell.value = user[i-64]
            pos += 1
        wb.save(file_name)
        wb.close()
        wb = None